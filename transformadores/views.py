from django.shortcuts import render, redirect
from .forms import InspeccionTransformadorForm
from .models import InspeccionTransformador
import openpyxl
from openpyxl.styles import Alignment
from django.http import HttpResponse
import os
from django.conf import settings

def menu_transformadores(request):
    return render(request, 'transformadores/menu_transformador.html')

def home(request):
    return render(request, 'transformadores/home.html')

def formulario_transformador(request):
    if request.method == 'POST':
        form = InspeccionTransformadorForm(request.POST)
        if form.is_valid():
            inspeccion = form.save()

            # Crear archivo Excel si no existe
            ruta_excel = os.path.join(settings.BASE_DIR, 'registro_inspecciones.xlsx')
            if os.path.exists(ruta_excel):
                wb = openpyxl.load_workbook(ruta_excel)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                # Encabezados solo si es nuevo
                ws.append([
                    'Fecha', 'Subestación', 'Tipo Transformador',
                    'Temperatura Devanado', 'Temperatura Aceite',
                    'Nivel Aceite Tanque', 'Nivel Aceite Conmutador',
                    'Nivel Aceite Bushings', 'Ventiladores', 'Óxido',
                    'Pintura', 'Vibración', 'Fugas de Aceite',
                    'Aisladores', 'Limpieza Transf.', 'Limpieza Rad.',
                    'Sílica Gel', 'Circuitos Control', 'Tierra',
                    'Taps', 'Tap Área Sombreada', 'Operaciones Tap',
                    'Pararrayos', 'Contador Pararrayos', 'Estado General',
                    'Observaciones'
                ])
            
            # Agregar fila
            ws.append([
                inspeccion.fecha, inspeccion.subestacion, inspeccion.tipo_transformador,
                inspeccion.temperatura_devanado, inspeccion.temperatura_aceite,
                inspeccion.nivel_aceite_tanque, inspeccion.nivel_aceite_conmutador,
                inspeccion.nivel_de_aceite_bushings, inspeccion.estado_funcionamiento_ventiladores,
                inspeccion.oxido_y_corrosión, inspeccion.pintura, inspeccion.vibracion_ruido,
                inspeccion.fugas_de_aceite, inspeccion.estado_de_aisladores,
                inspeccion.limpieza_de_transformadores, inspeccion.limpieza_de_radiadores,
                inspeccion.control_de_silica_gel, inspeccion.chequeo_de_circuitos_de_control,
                inspeccion.revisión_de_conexiones_a_tierra, inspeccion.posición_de_taps,
                inspeccion.indicador_de_cambio_de_tap_en_area_sombreada,
                inspeccion.numero_de_operaciones_del_cambiador_de_taps,
                inspeccion.estado_de_pararrayos, inspeccion.contador_de_descarga_de_pararrayos,
                inspeccion.estado_general, inspeccion.observaciones
            ])
            
            wb.save(ruta_excel)
            return redirect('registro_exitoso')
    else:
        form = InspeccionTransformadorForm()

    return render(request, 'transformadores/formulario_transformador.html', {'form': form})

def registro_exitoso(request):
    return render(request, 'transformadores/registro_exitoso.html')


def descargar_excel(request):
    ruta_excel = os.path.join(settings.BASE_DIR, 'registro_inspecciones.xlsx')
    if os.path.exists(ruta_excel):
        with open(ruta_excel, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=registro_inspecciones.xlsx'
            return response
    else:
        return HttpResponse("Archivo no encontrado", status=404)

def limpiar_datos_excel(request):
    ruta_excel = os.path.join(settings.BASE_DIR, 'registro_inspecciones.xlsx')
    if os.path.exists(ruta_excel):
        wb = openpyxl.load_workbook(ruta_excel)
        ws = wb.active

        # Eliminar todas las filas excepto la primera (encabezado)
        ws.delete_rows(2, ws.max_row)

        wb.save(ruta_excel)
        return HttpResponse("Datos del Excel eliminados (encabezados conservados).")
    else:
        return HttpResponse("Archivo no encontrado.", status=404)
